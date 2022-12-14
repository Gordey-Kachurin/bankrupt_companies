import pandas as pd
from openpyxl import load_workbook
import sys
import os
import shutil
import rarfile
import xlrd
import subprocess
from datetime import datetime

if os.getcwd() not in sys.path:
    sys.path.append(os.getcwd())
from settings import URLS, FOLDERS

from utils.exceptions import UnexpectedFileExtention


def rename_and_move(downloads_folder, temp_folder, region, year):

    for filename in os.listdir(temp_folder):
        new_name = region + " " + year + " "
        renamed = new_name + filename
        os.rename(
            os.path.join(temp_folder, filename),
            os.path.join(temp_folder, renamed),
        )
        try:
            shutil.move(
                os.path.join(temp_folder, renamed),
                os.path.join(downloads_folder, region),
            )
        except shutil.Error as e:
            print(e)
            os.remove(os.path.join(temp_folder, renamed))
            print(f"Removed: {renamed}")


def create_directories_for_copies(downloads_folder):
    for region in downloads_folder:
        if not os.path.exists(os.path.join(FOLDERS["copies"], region)):
            os.mkdir(os.path.join(FOLDERS["copies"], region))


def get_filenames_by_region(folder_with_regions):
    downloads = {}
    for region in URLS:
        try:
            files = os.listdir(os.path.join(folder_with_regions, region))
        except FileNotFoundError:
            continue
        downloads[region] = files

    return downloads


def get_regions_files_by_file_extention(downloads, file_extention):
    """
    file_extention examples ".xlsx", ".xls", ".rar"
    """
    regions_files_by_file_extention = {}
    for region in downloads:
        files = []
        for filename in downloads[region]:
            if filename.endswith(file_extention):
                files.append(filename)
        if files != []:
            regions_files_by_file_extention[region] = files
    return regions_files_by_file_extention


def extract_from_rar(regions_rar_files, use_7zip=True):
    """
    To use rarfile
        On Linux install unrar: sudo apt install unrar
        On windows download "UnRAR for Windows" - "Command line freeware Windows UnRAR."
            from https://www.rarlab.com/rar_add.htm
        Exctract file to current working directory
        Set rarfile.RarFile.UNRAR_TOOL to path to UnRAR.exe

    To use 7zip download 7zip: https://www.7-zip.org/
    """

    for region in regions_rar_files:
        for rar_filename in regions_rar_files[region]:
            splitted = rar_filename.split()
            year = splitted[1] + " " + splitted[2]

            if use_7zip:
                src_root = os.path.join(FOLDERS["downloads"], region, rar_filename)
                dest = FOLDERS["temp"]
                subprocess.run(f'7z x "{src_root}" -o"{dest}"')
                print(f"Extracted from: {rar_filename}")
            else:
                with rarfile.RarFile(
                    os.path.join(FOLDERS["downloads"], region, rar_filename)
                ) as rf:
                    rf.extractall(FOLDERS["temp"])
                    print(f"Extracted: {rf.filename}")

            rename_and_move(FOLDERS["downloads"], FOLDERS["temp"], region, year)


def delete_rar_files(region_rar_files):
    for region in region_rar_files:
        for rar_filename in region_rar_files[region]:
            os.remove(os.path.join(FOLDERS["downloads"], region, rar_filename))


def convert_xls_to_xlsx(
    path_to_source_file, path_to_destination_file, engine_for_source=None
):
    # TODO: Check to work with multiple sheets
    # xls = pd.ExcelFile('path_to_file.xls')
    # df1 = pd.read_excel(xls, 'Sheet1')
    # df2 = pd.read_excel(xls, 'Sheet2')
    pd.read_excel(path_to_source_file, engine=engine_for_source).to_excel(
        path_to_destination_file, engine="openpyxl", index=False
    )


def get_xlsx_content_headers(path_to_xlsx, patterns):
    wb = load_workbook(filename=path_to_xlsx)
    sheet = wb.active
    xlsx_cells = {}
    for field in patterns:
        for pattern in patterns[field]:
            for row in sheet.iter_rows():
                for cell in row:
                    try:
                        if pattern.match(cell.value):
                            print(f"{cell.coordinate}: {cell.value}")
                            xlsx_cells[field] = cell
                            break
                    # Handle cells with None value (empty cells)
                    except TypeError:
                        continue
                # When pattern not found in the current row,
                # look at the next row because next 'break'
                # will not allow you to look past first row
                if field not in xlsx_cells.keys():
                    continue
                # Check next field pattern
                break

    wb.close()
    return xlsx_cells


def copy_files(downloads, engine_for_source=None):
    """
    During copying all new line ('\n') characters in cells are removed.

    If used with muptiprocessing.Pool do not use print()

    Speed measurements using Pool from multiprocessing
    NO LOOPS
    No multiprocessing
        copy_files(regions_xls_files)
        copy_files(regions_xlsx_files)
    37.29260468482971 seconds

    Muptiprocessing
    Pool(4), 4 arguments: 28.903422594070435 seconds
    Pool(4), 2 arguments: 34.87284708023071 seconds
    Pool(2), 4 arguments: 30.20387887954712 seconds

    IN LOOPS
    range(10):
    p.map(copy_files, [regions_xls_files_part1, regions_xls_files_part2, regions_xlsx_files_part1, regions_xlsx_files_part2])
    Pool(8) 294.1417112350464 seconds
    Pool(4) 290.4357805252075 seconds
    Pool(2) 307.56513714790344 seconds

    p.map(copy_files, [regions_xls_files, regions_xlsx_files])
    Pool(2), 351.1649408340454 seconds
    Pool(4), 349.23773193359375 seconds
    Pool(8), 348.5626633167267 seconds

    No multiprocessing
    range(10):
        copy_files(regions_xls_files)
        copy_files(regions_xlsx_files)
    365.05810832977295 seconds


    Example of splitting 'downloads' dictionary argument for use in Pool:
    part = math.floor(len(regions_xls_files) / 2)
    regions_xls_files_part1 = dict(itertools.islice(regions_xls_files.items(), 0, part))
    regions_xls_files_part2 = dict(itertools.islice(regions_xls_files.items(), part, None))

    Check available CPUs to use as argument to Pool:
    len(os.sched_getaffinity(0))
    """
    for region in downloads:
        for filename in downloads[region]:
            print(filename)
            if filename.endswith(".xls"):
                new_filename = filename[:-3] + "xlsx"
                try:
                    convert_xls_to_xlsx(
                        os.path.join(FOLDERS["downloads"], region, filename),
                        os.path.join(FOLDERS["copies"], region, new_filename),
                        engine_for_source,
                    )
                except xlrd.XLRDError:
                    print(f"???????????????????????? ????????: {filename}")
                    continue

            elif filename.endswith(".xlsx"):
                convert_xls_to_xlsx(
                    os.path.join(FOLDERS["downloads"], region, filename),
                    os.path.join(FOLDERS["copies"], region, filename),
                )
            else:
                raise UnexpectedFileExtention


def prepare_excel_files_for_parsing():
    downloads = get_filenames_by_region(FOLDERS["downloads"])
    regions_rar_files = get_regions_files_by_file_extention(downloads, ".rar")
    extract_from_rar(regions_rar_files)
    delete_rar_files(regions_rar_files)
    create_directories_for_copies(downloads)
    regions_xls_files = get_regions_files_by_file_extention(downloads, ".xls")
    regions_xlsx_files = get_regions_files_by_file_extention(downloads, ".xlsx")
    copy_files(regions_xls_files, engine_for_source="xlrd")
    copy_files(regions_xlsx_files, engine_for_source=None)


def check_matched_patterns(xlsx_cells, patterns_dict):
    # Check for not matched patterns
    matched_fields = []
    not_matched_fields = []
    for field in patterns_dict:
        try:
            xlsx_cells[field]
            matched_fields.append(
                (field, xlsx_cells[field].coordinate, xlsx_cells[field].value)
            )
        except KeyError:
            not_matched_fields.append(field)
    return matched_fields, not_matched_fields


def map_fieds_to_filename(filename, not_matched_fields):
    not_matched_file = {}
    if not_matched_fields != []:
        not_matched_file[filename] = not_matched_fields
    return not_matched_file


def map_filename_to_region(region, not_matched_files):
    not_matched_region = {}
    if not_matched_files != []:
        not_matched_region[region] = not_matched_files
    return not_matched_region


def get_file_headers(patterns_dict):
    # TODO: Unexpected language in Excel file
    # https://akm.kgd.gov.kz/ru/content/informacionnye-soobshcheniya-4-3
    # Page in Russian, Excel file in Kazakh
    # Mangistau 2019 https://mng.kgd.gov.kz/ru/content/informacionnye-soobshcheniya-14-3
    # Unexpected filetype
    # ???????????????????? ?? ???????????????????? ???????????????? ????????????????????  ?? ?????????????????? ??????????????????????

    copies = get_filenames_by_region(FOLDERS["copies"])
    regions_xlsx_files = get_regions_files_by_file_extention(copies, ".xlsx")
    matched_files = []
    not_matched_files = []

    matched_region = {}
    not_matched_region = {}

    matched_to_return = []
    not_matched_to_return = []
    for region in regions_xlsx_files:

        for filename in regions_xlsx_files[region]:
            # print(filename)
            xlsx_cells = get_xlsx_content_headers(
                os.path.join(FOLDERS["copies"], region, filename), patterns_dict
            )
            matched_fields, not_matched_fields = check_matched_patterns(
                xlsx_cells, patterns_dict
            )
            matched_file = map_fieds_to_filename(filename, matched_fields.copy())
            not_matched_file = map_fieds_to_filename(
                filename, not_matched_fields.copy()
            )
            if matched_file != {}:
                matched_files.append(matched_file.copy())

            if not_matched_file != {}:
                not_matched_files.append(not_matched_file.copy())
        matched_region = map_filename_to_region(region, matched_files.copy())
        not_matched_region = map_filename_to_region(region, not_matched_files.copy())

        matched_to_return.append(matched_region.copy())
        not_matched_to_return.append(not_matched_region.copy())

        matched_files.clear()
        matched_region.clear()
        not_matched_files.clear()
        not_matched_region.clear()

    return matched_to_return, not_matched_to_return


def write_matched_or_not_dict_to_file(matched_or_not_dict, output_filename):
    now = datetime.now()
    current_time = now.strftime("%Y-%m-%d %H hour(s)")
    with open(os.path.join(FOLDERS["log"], output_filename), "a") as fp:
        for index, item in enumerate(matched_or_not_dict):
            for region in matched_or_not_dict[index]:
                for idx, fl in enumerate(item[region]):
                    for filename in item[region][idx]:
                        fp.write(
                            f"{current_time}: {filename}({item[region][idx][filename]})\n"
                        )
