import pandas as pd
from openpyxl import load_workbook
import sys
import os
import rarfile
import xlrd

if os.getcwd() not in sys.path:
    sys.path.append(os.getcwd())
from settings import (
    REGEX_PATTERNS_FOR_XLSX_HEADERS,
    URLS,
    FOLDERS,
)
from utils.parsers import rename_and_move
from utils.exceptions import UnexpectedFileExtention


def create_directories_for_copies(downloads_folder):
    for region in downloads_folder:
        if not os.path.exists(os.path.join(FOLDERS["copies"], region)):
            os.mkdir(os.path.join(FOLDERS["copies"], region))


def get_downloaded_filenames_by_region():
    downloads = {}
    for region in URLS:
        try:
            files = os.listdir(os.path.join(FOLDERS["downloads"], region))
        except FileNotFoundError:
            continue
        downloads[region] = files

    return downloads


def get_regions_files_by_file_extention(downloads, file_extention):
    """
    file_extention examples ".xlsx", ".xls", ".rar"
    """
    regions_files_by_file_extention = {}
    for region in downloads:
        files = []
        for filename in downloads[region]:
            if filename.endswith(file_extention):
                files.append(filename)
        if files != []:
            regions_files_by_file_extention[region] = files
    return regions_files_by_file_extention


# On Linux install unrar: sudo apt install unrar
def extract_rar_files(regions_rar_files):
    rar_file_names = []
    for region in regions_rar_files:
        for rar_filename in regions_rar_files[region]:
            splitted = rar_filename.split()
            year = splitted[1] + " " + splitted[2]
            rf = rarfile.RarFile(
                os.path.join(FOLDERS["downloads"], region, rar_filename)
            )
            print(rf.filename)
            rf.extractall(FOLDERS["temp"])
            rename_and_move(FOLDERS["downloads"], FOLDERS["temp"], region, year)


def delete_rar_files(region_rar_files):
    for region in region_rar_files:
        for rar_filename in region_rar_files[region]:
            os.remove(os.path.join(FOLDERS["downloads"], region, rar_filename))


def convert_xls_to_xlsx(
    path_to_source_file, path_to_destination_file, engine_for_source=None
):
    pd.read_excel(path_to_source_file, engine=engine_for_source).to_excel(
        path_to_destination_file, engine="openpyxl", index=False
    )


def get_xlsx_headers_cells(path_to_xlsx):
    wb = load_workbook(filename=path_to_xlsx)
    sheet = wb.active
    xlsx_headers = {}
    for row in sheet.iter_rows():
        for cell in row:
            for key in REGEX_PATTERNS_FOR_XLSX_HEADERS:
                try:
                    if REGEX_PATTERNS_FOR_XLSX_HEADERS[key].match(cell.value):
                        print(cell.coordinate, cell.value)
                        xlsx_headers[key] = cell
                except TypeError:
                    continue
    wb.close()
    return xlsx_headers


def copy_files(downloads, engine_for_source=None):
    """
    If used with muptiprocessing.Pool do not use print()

    """
    for region in downloads:
        for filename in downloads[region]:
            # print(region, filename)
            if filename.endswith(".xls"):
                new_filename = filename[:-3] + "xlsx"
                try:
                    convert_xls_to_xlsx(
                        os.path.join(FOLDERS["downloads"], region, filename),
                        os.path.join(FOLDERS["copies"], region, new_filename),
                        engine_for_source,
                    )
                except xlrd.XLRDError:
                    print(f"Поврежденный файл: {filename}")
                    continue

            elif filename.endswith(".xlsx"):
                convert_xls_to_xlsx(
                    os.path.join(FOLDERS["downloads"], region, filename),
                    os.path.join(FOLDERS["copies"], region, filename),
                )
            else:
                raise UnexpectedFileExtention


downloads = get_downloaded_filenames_by_region()
regions_rar_files = get_regions_files_by_file_extention(downloads, ".rar")
extract_rar_files(regions_rar_files)
delete_rar_files(regions_rar_files)
create_directories_for_copies(downloads)

regions_xls_files = get_regions_files_by_file_extention(downloads, ".xls")
regions_xlsx_files = get_regions_files_by_file_extention(downloads, ".xlsx")
regions_ods_files = get_regions_files_by_file_extention(downloads, ".ods")
regions_xlsm_files = get_regions_files_by_file_extention(downloads, ".xlsm")


import time
import math
from multiprocessing import Process, Pool
import itertools

part = math.floor(len(regions_xls_files) / 2)
regions_xls_files_part1 = dict(itertools.islice(regions_xls_files.items(), 0, part))
regions_xls_files_part2 = dict(itertools.islice(regions_xls_files.items(), part, None))

part_xlsx = math.floor(len(regions_xlsx_files) / 2)
regions_xlsx_files_part1 = dict(
    itertools.islice(regions_xlsx_files.items(), 0, part_xlsx)
)
regions_xlsx_files_part2 = dict(
    itertools.islice(regions_xlsx_files.items(), part_xlsx, None)
)


if __name__ == "__main__":
    start_time = time.time()

    with Pool(len(os.sched_getaffinity(0))) as p:
        p.map(
            copy_files,
            [
                regions_xls_files_part1,
                regions_xls_files_part2,
                regions_xlsx_files_part1,
                regions_xlsx_files_part2,
            ],
        )

    # XLS: range(10), Pool(1), 40.96643614768982 seconds
    # XLS: range(10), Pool(2), 27.535967111587524 seconds
    # XLS: range(10), Pool(3), 29.188135385513306 seconds
    # XLS: range(10), Pool(4), 27.81327795982361 seconds
    # XLS: range(10), Pool(5), 28.294697284698486 seconds
    # XLS: range(10), no pool, 42.82876753807068 seconds
    """
    Speed measurements using Pool from multiprocessing
    IN LOOPS
    range(10):
    p.map(copy_files, [regions_xls_files_part1, regions_xls_files_part2, regions_xlsx_files_part1, regions_xlsx_files_part2])
    Pool(8) 294.1417112350464 seconds
    Pool(4) 290.4357805252075 seconds
    Pool(2) 307.56513714790344 seconds

    p.map(copy_files, [regions_xls_files, regions_xlsx_files])
    Pool(2), 351.1649408340454 seconds
    Pool(4), 349.23773193359375 seconds
    Pool(8), 348.5626633167267 seconds

    No multiprocessing
    range(10):
        copy_files(regions_xls_files)
        copy_files(regions_xlsx_files)
    365.05810832977295 seconds

    ==============

    NO LOOPS
    No multiprocessing
        copy_files(regions_xls_files)
        copy_files(regions_xlsx_files)
    37.29260468482971 seconds

    Muptiprocessing
    Pool(4), 4 arguments: 28.903422594070435 seconds
    Pool(4), 2 arguments: 34.87284708023071 seconds
    Pool(2), 4 arguments: 30.20387887954712 seconds
    """

    print("--- %s seconds ---" % (time.time() - start_time))


# for region in downloads:
#     print(downloads[region])
#     for f in downloads[region]:
#         convert_xls_to_xlsx(os.path.join(ROOT_FOLDER, region, f), os.path.join(COPIES_FOLDER, region, f))
